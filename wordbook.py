# coding=gbk

import pandas as pd
from pandas import DataFrame
from tkinter import *
import tkinter.font as tkFont
import random
import datetime
import time
import pygame
from openpyxl import load_workbook

wb = pd.read_excel('wordbook.xlsx')
all_word = wb.to_dict(orient='records')

reviewlist = []

#艾宾浩斯遗忘曲线
n = 0
while n < len(all_word):
	if all_word[n]['Progress'] == 0:
		reviewlist.append(all_word[n])
		del all_word[n]
		n -= 1
	elif all_word[n]['Progress'] == 1:
		if (all_word[n]['LastReview'] + datetime.timedelta(minutes=5)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 2:
		if (all_word[n]['LastReview'] + datetime.timedelta(minutes=30)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 3:
		if (all_word[n]['LastReview'] + datetime.timedelta(hours=12)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 4:
		if (all_word[n]['LastReview'] + datetime.timedelta(days=1)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 5:
		if (all_word[n]['LastReview'] + datetime.timedelta(days=2)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 6:
		if (all_word[n]['LastReview'] + datetime.timedelta(days=4)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 7:
		if (all_word[n]['LastReview'] + datetime.timedelta(days=7)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 8:
		if (all_word[n]['LastReview'] + datetime.timedelta(days=14)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 9:
		if (all_word[n]['LastReview'] + datetime.timedelta(weeks=4)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 10:
		if (all_word[n]['LastReview'] + datetime.timedelta(weeks=8)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 11:
		if (all_word[n]['LastReview'] + datetime.timedelta(weeks=12)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	elif all_word[n]['Progress'] == 12:
		if (all_word[n]['LastReview'] + datetime.timedelta(weeks=16)) < datetime.datetime.now():
			reviewlist.append(all_word[n])
			del all_word[n]
			n -= 1
	n +=1
random.shuffle(reviewlist)

def showword():
	english.delete('1.0','end')
	chinese.delete('1.0','end')
	i = random.choice([0,1])
	#i = 0
	if i == 0:
		english.insert('insert', currentreview[0]['English'])
		pygame.mixer.init()
		track = pygame.mixer.music.load("audio/" + currentreview[0]['English'] + ".mp3")
		pygame.mixer.music.play()
	if i == 1:
		chinese.insert('insert', currentreview[0]['Chinese'])
		hint['state'] = 'normal'
	answer['state'] = 'normal'
	remember_button['state'] = 'disabled'
	forget_button['state'] = 'disabled'
	play['state'] = 'disabled'
	answer.focus_set()

root = Tk()
root.title('耿浩翔单词本')
root.geometry('+300+15')

need_review_number = len(reviewlist)
review_this_time = 0
test = 0
currentreview = []

Label(root, text="需要复习").grid(row=1, column=1, padx=10, pady=10)
need_review = Entry(root, width=8)
need_review.insert(END, need_review_number)
need_review.grid(row=1, column=2, padx=10, pady=10)
review_amount = Entry(root, width=8)
review_amount.insert(END, need_review_number)
def eventhandler2(event):
	review()
review_amount.bind('<Return>', eventhandler2)
def eventhandler3(event):
	global review_this_time
	review_this_time = int(review_amount.get())
	if review_this_time > 10:
		review_amount.delete(0, END)
		review_amount.insert(END, 10)
review_amount.bind('<KeyPress-Down>', eventhandler3)
review_amount.grid(row=1, column=3, padx=10, pady=10)
ft1 = tkFont.Font(size=25)
english = Text(root, width=32, height=3, font=ft1)
english.insert(END, '')
english.grid(row=2, column=1, columnspan=3, padx=10, pady=10)
def pronounce():
	pygame.mixer.init()
	track = pygame.mixer.music.load("audio/" + currentreview[0]['English'] + ".mp3")
	pygame.mixer.music.play()
play = Button(root, width=7, height=4, text="播放", state=DISABLED, command=pronounce)
play.grid(row=2, column=4, columnspan=3, padx=10, pady=10)

chinese = Text(root, width=32, height=10, font=ft1)
chinese.insert(END, '')
chinese.grid(row=3, column=1, rowspan=3, columnspan=3, padx=0, pady=10)

def show_firstletter():
	english.insert('insert', currentreview[0]['English'][0:1])
	hint['state'] = 'disabled'
hint = Button(root, width=7, height=2, text="提示", state=DISABLED, command=show_firstletter)
hint.grid(row=3, column=4, columnspan=4, padx=10, pady=10)

def show_answer():
	english.delete('1.0','end')
	chinese.delete('1.0','end') 
	english.insert('insert', currentreview[0]['English'])
	chinese.insert('insert', currentreview[0]['Chinese'])
	pygame.mixer.init()
	track = pygame.mixer.music.load("audio/" + currentreview[0]['English'] + ".mp3")
	pygame.mixer.music.play()
	answer['state'] = 'disabled'
	hint['state'] = 'disabled'
	remember_button['state'] = 'normal'
	forget_button['state'] = 'normal'
	play['state'] = 'normal'
	undo_button['state'] = 'disabled'
	remember_button.focus_set()
answer = Button(root, height=12, text="显示答案", state=DISABLED, command=show_answer)
def eventhandler5(event):
	if hint['state'] == 'normal':
		show_firstletter()
answer.bind('<KeyPress-Up>', eventhandler5 )
answer.grid(row=4, column=4, rowspan=2, columnspan=4, padx=10, pady=10)

def review():
	global review_this_time
	global test
	review_this_time = int(review_amount.get())
	test = int(review_amount.get())
	n = 0
	while n < review_this_time:
		currentreview.append(reviewlist[n])
		del reviewlist[n]
		review_this_time -= 1
	review_this_time = int(review_amount.get())
	start_review['state'] = 'disabled'
	sync_button['state'] = 'disabled'
	showword()
start_review = Button(root, text="开始复习", command=review)
def eventhandler4(event):
	review_amount.focus_set()
start_review.bind('<KeyPress-Left>', eventhandler4)
start_review.grid(row=1, column=4, padx=10, pady=10)
start_review.focus_set()

def remember():
	global review_this_time
	global test
	if test > 0:
		currentreview[0]['Progress'] += 1
		currentreview[0]['LastReview'] = datetime.datetime.now()
		all_word.append(currentreview[0])
	test -= 1
	del currentreview[0]
	review_this_time -= 1
	review_amount.delete(0, END)
	review_amount.insert('insert', str(review_this_time))
	english.delete('1.0','end')
	chinese.delete('1.0','end')
	answer.focus_set()
	if review_this_time != 0:
		showword()
		if test >= 0:
			undo_button['state'] = 'normal'
	else:
		if len(reviewlist) == 0:
			need_review.delete(0, END)
			need_review.insert('insert', str(len(reviewlist)))
			random.shuffle(all_word)
			try:
				data=DataFrame(all_word)
				data = data[['English','Chinese','Progress','LastReview']]
				data.to_excel('wordbook.xlsx')
				wb = load_workbook('wordbook.xlsx')
				ws = wb.active
				ws.column_dimensions['B'].width = 25
				ws.column_dimensions['C'].width = 65
				ws.column_dimensions['D'].width = 15
				ws.column_dimensions['E'].width = 25
				wb.save('wordbook.xlsx')
				english.delete('1.0','end')
				english.insert('insert', "已完成所有复习！")
				quit_button.focus_set()
			except:
				english.insert('insert', "请关闭文档")
				sync_button.focus_set()
		else:
			need_review.delete(0, END)
			need_review.insert('insert', str(len(reviewlist)))
			review_amount.delete(0, END)
			review_amount.insert('insert', str(len(reviewlist)))
			start_review['state'] = 'normal'
			english.delete('1.0','end')
			english.insert('insert', "请同步后再关闭文档！")
			sync_button.focus_set()
		play['state'] = 'disabled'
		remember_button['state'] = 'disabled'
		forget_button['state'] = 'disabled'
		sync_button['state'] = 'normal'
remember_button = Button(root, width=16, text="记得", state=DISABLED, command=remember)
def eventhandler1(event):
	forget()
remember_button.bind('<KeyPress-Right>', eventhandler1 )
remember_button.grid(row=6, column=1, rowspan=2, columnspan=2, padx=10, pady=10)

def forget():
	global review_this_time
	global test
	if test > 0:
		if currentreview[0]['Progress'] != 0:
			currentreview[0]['Progress'] -= 1
			currentreview[0]['LastReview'] = datetime.datetime.now()
		all_word.append(currentreview[0])
	test -= 1
	currentreview.append(currentreview[0])
	del currentreview[0]
	english.delete('1.0','end')
	chinese.delete('1.0','end') 
	showword()
forget_button = Button(root, width=16, text="忘记", state=DISABLED, command=forget)
forget_button.grid(row=6, column=3, rowspan=2, columnspan=2, padx=10, pady=10)

def undo():
	global review_this_time
	global test
	test += 1
	review_this_time += 1
	review_amount.delete(0, END)
	review_amount.insert('insert', str(review_this_time))
	currentreview.insert(0, all_word[-1])
	currentreview[0]['Progress'] -= 1
	del all_word[-1]
	showword()
	undo_button['state'] = 'disabled'
undo_button = Button(root, width=7, text="撤销", state=DISABLED, command=undo)
undo_button.grid(row=6, column=4, rowspan=2, padx=10, pady=10)

def synchonize():
	reviewed=DataFrame(all_word)
	notreviewed=DataFrame(reviewlist)
	try:
		data = notreviewed.append(reviewed)
		data = data[['English','Chinese','Progress','LastReview']]
		data.to_excel('wordbook.xlsx')
		wb = load_workbook('wordbook.xlsx')
		ws = wb.active
		ws.column_dimensions['B'].width = 25
		ws.column_dimensions['C'].width = 65
		ws.column_dimensions['D'].width = 15
		ws.column_dimensions['E'].width = 25
		wb.save('wordbook.xlsx')
		english.delete('1.0','end')
		english.insert('insert', "同步完成")
		review_amount.focus_set()
	except:
		english.insert('insert', "请关闭文档")
sync_button = Button(root, width=16, text="同步", state=DISABLED, command=synchonize)
sync_button.grid(row=8, column=1, rowspan=1, padx=10, pady=10)

def quitprogram():
	root.destroy()
quit_button = Button(root, width=16, text="退出", command=quitprogram)
quit_button.grid(row=8, column=3, rowspan=1, padx=10, pady=10)

root.mainloop()
